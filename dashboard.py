"""
APPS — Purchase Planning Dashboard (Responsive)
รันด้วย: python3 -m streamlit run dashboard.py
รองรับทั้ง Desktop และ Mobile — full function เหมือนกัน
"""
import io
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
from pathlib import Path
from datetime import datetime

EXCEL_PATH = Path(__file__).parent / "output" / "purchasing_plan.xlsx"
SHEET_ALL  = "ทั้งหมด"

st.set_page_config(
    page_title="แผนสั่งซื้อ | AIMAN",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="auto",
)

# ── Device detection via JS ────────────────────────────────────────
# JS เขียน viewport width ลง query param ?w=xxx แล้ว Python อ่าน
if "device_detected" not in st.session_state:
    st.session_state.device_detected = False

components.html("""
<script>
(function() {
    const w = window.innerWidth;
    const url = new URL(window.parent.location.href);
    const current = url.searchParams.get('w');
    if (!current || Math.abs(parseInt(current) - w) > 100) {
        url.searchParams.set('w', w);
        window.parent.history.replaceState({}, '', url.toString());
    }
})();
</script>
""", height=0)

_w = st.query_params.get("w", "1200")
try:
    viewport_w = int(_w)
except ValueError:
    viewport_w = 1200

IS_MOBILE = viewport_w < 768
DEVICE_LABEL = "📱 มือถือ" if IS_MOBILE else "🖥️ คอมพิวเตอร์"

# ── Global CSS ─────────────────────────────────────────────────────
_mob_hide = """
    #MainMenu, footer, header { display: none !important; }
    .block-container { padding: 8px 12px !important; max-width: 100% !important; }
    [data-testid="stSidebar"] { display: none !important; }
""" if IS_MOBILE else ""

_block_padding = "padding: 12px 16px !important; max-width: 100% !important;" if IS_MOBILE else "padding: 24px 32px !important; max-width: 1400px !important;"
_mob_btn = """
    div[data-testid="stButton"] button { font-size: 14px !important; padding: 10px 12px !important; min-height: 44px !important; }
    div[data-testid="stNumberInput"] input { font-size: 16px !important; min-height: 44px !important; }
""" if IS_MOBILE else ""

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sarabun:wght@300;400;600;700;800&display=swap');

*, *::before, *::after {{
    font-family: 'Sarabun', sans-serif !important;
    box-sizing: border-box;
}}

/* ── hide streamlit chrome on mobile ── */
{_mob_hide}

/* ── block container ── */
.block-container {{
    {_block_padding}
}}

/* ── TOP BAR ── */
.top-bar {{
    background: linear-gradient(135deg, #0f2444 0%, #1a3a6e 100%);
    color: white;
    border-radius: {"10px" if IS_MOBILE else "16px"};
    padding: {"14px 16px" if IS_MOBILE else "20px 28px"};
    margin-bottom: {"12px" if IS_MOBILE else "20px"};
    display: flex;
    justify-content: space-between;
    align-items: center;
    gap: 12px;
}}
.top-bar h1 {{
    margin: 0;
    font-size: {"17px" if IS_MOBILE else "24px"};
    font-weight: 800;
    line-height: 1.2;
}}
.top-bar .device-tag {{
    font-size: {"10px" if IS_MOBILE else "11px"};
    background: rgba(255,255,255,0.15);
    border-radius: 20px;
    padding: 2px 10px;
    margin-top: 4px;
    display: inline-block;
}}
.top-bar .updated {{ font-size: {"10px" if IS_MOBILE else "12px"}; color: #8fb0d8; margin-top: 2px; }}
.top-bar .total-block {{ text-align: right; flex-shrink: 0; }}
.top-bar .total-lbl {{ font-size: {"10px" if IS_MOBILE else "11px"}; color: #8fb0d8; }}
.top-bar .total-num {{ font-size: {"20px" if IS_MOBILE else "28px"}; font-weight: 800; color: #ffd700; line-height: 1.1; }}
.top-bar .total-cnt {{ font-size: {"10px" if IS_MOBILE else "11px"}; color: #aac; }}

/* ── STAT STRIP ── */
.stat-strip {{
    display: flex;
    gap: {"6px" if IS_MOBILE else "12px"};
    margin-bottom: {"10px" if IS_MOBILE else "20px"};
    overflow-x: auto;
}}
.stat-box {{
    flex: 1;
    min-width: {"58px" if IS_MOBILE else "90px"};
    border-radius: {"8px" if IS_MOBILE else "12px"};
    padding: {"10px 8px" if IS_MOBILE else "16px 12px"};
    text-align: center;
    border: 1.5px solid rgba(0,0,0,0.08);
}}
.stat-box .s-num {{ font-size: {"22px" if IS_MOBILE else "32px"}; font-weight: 800; line-height: 1; }}
.stat-box .s-lbl {{ font-size: {"10px" if IS_MOBILE else "12px"}; font-weight: 600; margin-top: 4px; }}

/* ── SECTION TITLE ── */
.section-title {{
    font-size: {"14px" if IS_MOBILE else "17px"};
    font-weight: 700;
    color: #0f2444;
    border-left: 4px solid #e74c3c;
    padding-left: 10px;
    margin: {"14px 0 8px 0" if IS_MOBILE else "22px 0 12px 0"};
}}

/* ── PRODUCT CARD ── */
.pcard {{
    background: #fff;
    border-radius: {"10px" if IS_MOBILE else "12px"};
    padding: {"10px 12px" if IS_MOBILE else "14px 18px"};
    margin-bottom: {"6px" if IS_MOBILE else "8px"};
    border-left: 5px solid #e74c3c;
    box-shadow: 0 1px 6px rgba(0,0,0,0.07);
    transition: box-shadow .15s;
}}
.pcard.warn   {{ border-left-color: #f39c12; }}
.pcard.ok     {{ border-left-color: #27ae60; }}
.pcard.over   {{ border-left-color: #95a5a6; }}
.pcard.conf   {{ border-left-color: #27ae60; background: #f6fff9; }}

.pcard-top {{ display: flex; align-items: flex-start; gap: {"8px" if IS_MOBILE else "14px"}; }}
.pname {{
    flex: 1;
    font-size: {"13px" if IS_MOBILE else "15px"};
    font-weight: 700;
    color: #0f2444;
    line-height: 1.3;
}}
.pmeta {{ font-size: {"10px" if IS_MOBILE else "11px"}; color: #999; margin-top: 2px; }}
.psup  {{ font-size: {"10px" if IS_MOBILE else "11px"}; color: #557; margin-top: 1px; }}
.pnote {{ font-size: {"10px" if IS_MOBILE else "11px"}; color: #c0392b; margin-top: 2px; }}

.kpis {{ display: flex; gap: {"8px" if IS_MOBILE else "12px"}; align-items: center; flex-shrink: 0; }}
.kpi  {{ text-align: center; }}
.kpi-num {{ font-size: {"16px" if IS_MOBILE else "20px"}; font-weight: 800; line-height: 1; }}
.kpi-lbl {{ font-size: {"9px" if IS_MOBILE else "10px"}; color: #aaa; margin-top: 2px; }}

.qty-chip {{
    background: #fff8e1;
    border-radius: {"7px" if IS_MOBILE else "8px"};
    padding: {"6px 10px" if IS_MOBILE else "8px 14px"};
    text-align: center;
    flex-shrink: 0;
}}
.pcard.conf .qty-chip {{ background: #e8f8ee; }}
.qty-chip-num {{ font-size: {"15px" if IS_MOBILE else "18px"}; font-weight: 800; color: #0f2444; }}
.qty-chip-lbl {{ font-size: {"9px" if IS_MOBILE else "10px"}; color: #888; }}

/* ── SALES ROW ── */
.sales-row {{
    display: flex;
    margin-top: {"8px" if IS_MOBILE else "10px"};
    border-top: 1px solid #f3f3f3;
    padding-top: {"6px" if IS_MOBILE else "8px"};
    gap: 0;
}}
.sc {{
    flex: 1;
    text-align: center;
    border-right: 1px solid #f0f0f0;
    padding: 0 2px;
}}
.sc:last-child {{ border-right: none; }}
.sc-num {{ font-size: {"13px" if IS_MOBILE else "15px"}; font-weight: 700; color: #333; }}
.sc-lbl {{ font-size: {"9px" if IS_MOBILE else "10px"}; color: #bbb; margin-top: 1px; }}

/* ── CONFIRMED BANNER ── */
.conf-banner {{
    background: linear-gradient(135deg, #155724, #27ae60);
    color: white;
    border-radius: {"10px" if IS_MOBILE else "14px"};
    padding: {"14px 16px" if IS_MOBILE else "18px 24px"};
    margin: {"12px 0" if IS_MOBILE else "20px 0"};
    display: flex;
    justify-content: space-between;
    align-items: center;
    gap: 12px;
}}
.cb-title {{ font-size: {"14px" if IS_MOBILE else "17px"}; font-weight: 800; }}
.cb-sub   {{ font-size: {"11px" if IS_MOBILE else "12px"}; color: #b2f0c8; margin-top: 3px; }}
.cb-total {{ font-size: {"22px" if IS_MOBILE else "30px"}; font-weight: 800; color: #ffd700; }}
.cb-lbl   {{ font-size: {"10px" if IS_MOBILE else "11px"}; color: #b2f0c8; }}

/* ── MOBILE FILTER BAR ── */
.mob-filter-wrap {{
    background: #f8f9fb;
    border-radius: 10px;
    padding: 12px;
    margin-bottom: 10px;
    border: 1px solid #e8eaf0;
}}

/* ── SUPPLIER GROUP HEADER ── */
.sup-header {{
    font-size: {"12px" if IS_MOBILE else "14px"};
    font-weight: 700;
    color: #1a3a6e;
    background: #eef3ff;
    border-radius: 8px;
    padding: {"6px 10px" if IS_MOBILE else "8px 14px"};
    margin: {"8px 0 4px 0" if IS_MOBILE else "12px 0 6px 0"};
    display: flex;
    justify-content: space-between;
    align-items: center;
}}

/* ── NO ORDER ── */
.no-order {{
    text-align: center;
    padding: {"30px 16px" if IS_MOBILE else "60px"};
    color: #27ae60;
    font-size: {"15px" if IS_MOBILE else "20px"};
    font-weight: 700;
}}

/* ── Streamlit button overrides ── */
{_mob_btn}
</style>
""", unsafe_allow_html=True)


# ── Session state ─────────────────────────────────────────────────
if "confirmed" not in st.session_state:
    st.session_state.confirmed = {}

# ── Load data ─────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def load_data():
    df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_ALL, engine="openpyxl")
    num_cols = [
        "คงเหลือ", "PO ค้างรับ", "Forecast/วัน", "Days of Supply",
        "Safety Stock", "Reorder Point", "แนะนำสั่ง (หน่วย)", "มูลค่าสั่ง (บาท)",
        "ราคาต้นทุน (บาท)",
        "ขาย 3 วัน", "ขาย 7 วัน", "ขาย 15 วัน",
        "ขาย 30 วัน", "ขาย 60 วัน", "ขาย 90 วัน",
    ]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    return df

if not EXCEL_PATH.exists():
    st.error("❌ ไม่พบไฟล์ข้อมูล — รัน `python3 run_planning.py` ก่อน")
    st.stop()

df    = load_data()
mtime = datetime.fromtimestamp(EXCEL_PATH.stat().st_mtime)

need_order = df[df["Alert"].isin(["CRITICAL", "WARNING"])].copy()
critical   = df[df["Alert"] == "CRITICAL"]
warning    = df[df["Alert"] == "WARNING"]
overstock  = df[df["Alert"] == "OVERSTOCK"]
ok_df      = df[df["Alert"] == "OK"]

total_order_value = need_order["มูลค่าสั่ง (บาท)"].sum() if "มูลค่าสั่ง (บาท)" in need_order.columns else 0
confirmed_total   = sum(v["qty"] * v["unit_cost"] for v in st.session_state.confirmed.values())
confirmed_count   = len(st.session_state.confirmed)

# ── TOP BAR ──────────────────────────────────────────────────────
st.markdown(f"""
<div class="top-bar">
  <div>
    <h1>📦 แผนสั่งซื้อสินค้า</h1>
    <span class="device-tag">{DEVICE_LABEL} · {viewport_w}px</span>
    <div class="updated">อัปเดต {mtime.strftime('%d %b %Y เวลา %H:%M น.')}</div>
  </div>
  <div class="total-block">
    <div class="total-lbl">มูลค่าที่ควรสั่ง</div>
    <div class="total-num">฿{total_order_value:,.0f}</div>
    <div class="total-cnt">{len(need_order)} รายการ</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── STAT STRIP ────────────────────────────────────────────────────
st.markdown(f"""
<div class="stat-strip">
  <div class="stat-box" style="background:#fff0f0;border-color:#ffc0c0">
    <div class="s-num" style="color:#e74c3c">{len(critical)}</div>
    <div class="s-lbl">🔴 ด่วน</div>
  </div>
  <div class="stat-box" style="background:#fff8f0;border-color:#ffd9a0">
    <div class="s-num" style="color:#e67e22">{len(warning)}</div>
    <div class="s-lbl">🟠 เร่งด่วน</div>
  </div>
  <div class="stat-box" style="background:#f0fff4;border-color:#a0e0b0">
    <div class="s-num" style="color:#27ae60">{confirmed_count}</div>
    <div class="s-lbl">✅ Confirmed</div>
  </div>
  <div class="stat-box" style="background:#f0f8ff;border-color:#a0c8e0">
    <div class="s-num" style="color:#2980b9">{len(ok_df)}</div>
    <div class="s-lbl">🟢 ปกติ</div>
  </div>
  <div class="stat-box" style="background:#f5f5f5;border-color:#ccc">
    <div class="s-num" style="color:#888">{len(overstock)}</div>
    <div class="s-lbl">⚪ เกิน</div>
  </div>
</div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
#  FILTERS — Sidebar (Desktop) / Expander (Mobile)
# ══════════════════════════════════════════════════════════════════
suppliers = ["ทั้งหมด"] + sorted(df["Supplier"].dropna().astype(str).unique().tolist())

def _render_filters(in_sidebar=False):
    prefix = "sb_" if in_sidebar else "mb_"

    col_r, col_c = (st.columns([1,1]) if in_sidebar else (st.container(), st.container()))
    if in_sidebar:
        with col_r:
            if st.button("🔄 โหลดใหม่", use_container_width=True, key=f"{prefix}reload"):
                st.cache_data.clear()
                st.session_state.confirmed = {}  # clear confirmed เมื่อโหลดข้อมูลใหม่
                st.rerun()
        with col_c:
            if st.button("🗑️ ล้าง Confirm", use_container_width=True, key=f"{prefix}clear", type="secondary"):
                st.session_state.confirmed = {}
                st.rerun()
    else:
        c1, c2 = st.columns(2)
        with c1:
            if st.button("🔄 โหลดใหม่", use_container_width=True, key=f"{prefix}reload"):
                st.cache_data.clear()
                st.session_state.confirmed = {}  # clear confirmed เมื่อโหลดข้อมูลใหม่
                st.rerun()
        with c2:
            if st.button("🗑️ ล้าง Confirm", use_container_width=True, key=f"{prefix}clear", type="secondary"):
                st.session_state.confirmed = {}
                st.rerun()

    show_levels = st.multiselect(
        "แสดงระดับ",
        ["CRITICAL", "WARNING", "OK", "OVERSTOCK"],
        default=["CRITICAL", "WARNING"],
        key=f"{prefix}levels",
    )
    sel_sup = st.selectbox("Supplier", suppliers, key=f"{prefix}sup")
    search  = st.text_input("🔎 ค้นหาชื่อ / SKU", key=f"{prefix}search")
    group   = st.checkbox("จัดกลุ่มตาม Supplier", value=True, key=f"{prefix}group")

    if in_sidebar:
        st.markdown("---")
        st.caption(f"**{DEVICE_LABEL}** · viewport {viewport_w}px")
        st.markdown("""
**วิธีใช้**
- กด **✅ Confirm** เพื่อเพิ่มสินค้าในรายการสั่ง
- แก้จำนวนที่ช่อง input ก่อนกด Confirm
- กด **📥 Export PO** เพื่อโหลด Excel แยกซัพลายเยอร์
""")

    return show_levels, sel_sup, search, group


if IS_MOBILE:
    with st.expander("⚙️ ตัวกรองและการตั้งค่า", expanded=False):
        show_levels, sel_sup, search, group_by_sup = _render_filters(in_sidebar=False)
else:
    with st.sidebar:
        st.markdown("### ⚙️ ตัวกรอง")
        show_levels, sel_sup, search, group_by_sup = _render_filters(in_sidebar=True)


# ── FILTER + SORT DATA ────────────────────────────────────────────
filtered = df[df["Alert"].isin(show_levels)].copy()
if sel_sup != "ทั้งหมด":
    filtered = filtered[filtered["Supplier"].astype(str) == sel_sup]
if search:
    mask = (
        filtered["ชื่อสินค้า"].astype(str).str.contains(search, case=False, na=False) |
        filtered["SKU ID"].astype(str).str.contains(search, case=False, na=False)
    )
    filtered = filtered[mask]

level_ord = {"CRITICAL": 0, "WARNING": 1, "WATCH": 2, "OK": 3, "OVERSTOCK": 4}
filtered["_ord"] = filtered["Alert"].map(level_ord)
filtered = filtered.sort_values(["_ord", "Days of Supply"]).drop(columns=["_ord"])

order_rows = filtered[filtered["Alert"].isin(["CRITICAL", "WARNING"])]
other_rows = filtered[~filtered["Alert"].isin(["CRITICAL", "WARNING"])]


# ══════════════════════════════════════════════════════════════════
#  RENDER PRODUCT CARD
# ══════════════════════════════════════════════════════════════════
def render_item(row):
    sku_id      = str(row.get("SKU ID", ""))
    sku_name    = str(row.get("ชื่อสินค้า", ""))
    supplier    = str(row.get("Supplier", ""))
    alert       = str(row.get("Alert", ""))
    dos         = float(row.get("Days of Supply", 0) or 0)
    stock       = int(row.get("คงเหลือ", 0) or 0)
    default_qty = int(row.get("แนะนำสั่ง (หน่วย)", 0) or 0)
    unit_cost   = float(row.get("ราคาต้นทุน (บาท)", 0) or 0)
    note        = str(row.get("หมายเหตุ", "") or "")
    s3  = int(row.get("ขาย 3 วัน",  0) or 0)
    s7  = int(row.get("ขาย 7 วัน",  0) or 0)
    s15 = int(row.get("ขาย 15 วัน", 0) or 0)
    s30 = int(row.get("ขาย 30 วัน", 0) or 0)
    s60 = int(row.get("ขาย 60 วัน", 0) or 0)
    s90 = int(row.get("ขาย 90 วัน", 0) or 0)

    is_confirmed  = sku_id in st.session_state.confirmed
    confirmed_qty = st.session_state.confirmed.get(sku_id, {}).get("qty", default_qty)

    card_cls  = "pcard conf" if is_confirmed else (
                "pcard warn" if alert == "WARNING" else "pcard")
    dos_color = ("#27ae60" if is_confirmed else
                 "#f39c12" if alert == "WARNING" else "#e74c3c")
    note_html = f'<div class="pnote">⚠ {note[:80]}</div>' if note else ""

    st.markdown(f"""
<div class="{card_cls}">
  <div class="pcard-top">
    <div style="flex:1;min-width:0">
      <div class="pname">{sku_name}</div>
      <div class="pmeta">SKU: {sku_id}</div>
      <div class="psup">🏭 {supplier}</div>
      {note_html}
    </div>
    <div class="kpis">
      <div class="kpi">
        <div class="kpi-num" style="color:{dos_color}">{dos:.0f}</div>
        <div class="kpi-lbl">วันเหลือ</div>
      </div>
      <div class="kpi">
        <div class="kpi-num" style="color:#2980b9">{stock:,}</div>
        <div class="kpi-lbl">คงเหลือ</div>
      </div>
      <div class="qty-chip">
        <div class="qty-chip-num">{confirmed_qty:,}</div>
        <div class="qty-chip-lbl">{"✅ confirmed" if is_confirmed else "แนะนำสั่ง"}</div>
      </div>
    </div>
  </div>
  <div class="sales-row">
    <div class="sc"><div class="sc-num">{s3:,}</div><div class="sc-lbl">3ว</div></div>
    <div class="sc"><div class="sc-num">{s7:,}</div><div class="sc-lbl">7ว</div></div>
    <div class="sc"><div class="sc-num">{s15:,}</div><div class="sc-lbl">15ว</div></div>
    <div class="sc"><div class="sc-num">{s30:,}</div><div class="sc-lbl">30ว</div></div>
    <div class="sc"><div class="sc-num">{s60:,}</div><div class="sc-lbl">60ว</div></div>
    <div class="sc"><div class="sc-num">{s90:,}</div><div class="sc-lbl">90ว</div></div>
    <div class="sc"><div class="sc-num" style="color:#27ae60">฿{confirmed_qty*unit_cost:,.0f}</div><div class="sc-lbl">มูลค่า</div></div>
  </div>
</div>
""", unsafe_allow_html=True)

    # Quantity input — right-aligned with visible label
    _spacer, _inp_col = st.columns([2, 1])
    with _inp_col:
        st.number_input(
            "จำนวนที่สั่งจริง", min_value=0, value=confirmed_qty,
            step=1, key=f"qty_{sku_id}"
        )


# ══════════════════════════════════════════════════════════════════
#  MAIN — รายการต้องสั่ง
# ══════════════════════════════════════════════════════════════════
if len(order_rows) == 0:
    st.markdown('<div class="no-order">✅ ไม่มีสินค้าที่ต้องสั่งในขณะนี้</div>',
                unsafe_allow_html=True)
else:
    st.markdown(f'<div class="section-title">⚠️ ต้องสั่งซื้อ — {len(order_rows)} รายการ</div>',
                unsafe_allow_html=True)

    if group_by_sup:
        for sup, grp in order_rows.groupby("Supplier", sort=False):
            sup_val = grp["มูลค่าสั่ง (บาท)"].sum()
            sup_skus = grp["SKU ID"].astype(str).tolist()
            all_confirmed = all(s in st.session_state.confirmed for s in sup_skus)
            any_confirmed = any(s in st.session_state.confirmed for s in sup_skus)
            conf_badge = " ✅" if all_confirmed else (" ◑" if any_confirmed else "")
            fsize = '11px' if IS_MOBILE else '12px'
            st.markdown(f"""
<div class="sup-header">
  <span>🏭 {sup}{conf_badge}</span>
  <span style="color:#557;font-size:{fsize}">
    {len(grp)} รายการ · แนะนำ ฿{sup_val:,.0f}
  </span>
</div>""", unsafe_allow_html=True)

            for _, row in grp.iterrows():
                render_item(row)

            # ── Per-supplier confirm / cancel buttons ──────────────
            sup_key = sup.replace(" ", "_").replace("/", "-")
            if IS_MOBILE:
                if any_confirmed:
                    if st.button(f"❌ ยกเลิกทั้งหมด ({sup})",
                                 key=f"sup_cancel_{sup_key}", use_container_width=True):
                        for s in sup_skus:
                            st.session_state.confirmed.pop(s, None)
                        st.rerun()
                if st.button(f"✅ Confirm สั่งทั้งหมด — {sup} ({len(grp)} รายการ)",
                             key=f"sup_confirm_{sup_key}",
                             use_container_width=True, type="primary"):
                    for _, row in grp.iterrows():
                        s = str(row["SKU ID"])
                        qty_val = int(st.session_state.get(f"qty_{s}",
                                      int(row.get("แนะนำสั่ง (หน่วย)", 0) or 0)))
                        st.session_state.confirmed[s] = {
                            "sku_id": s,
                            "sku_name": str(row.get("ชื่อสินค้า", "")),
                            "supplier": sup,
                            "qty": qty_val,
                            "unit_cost": float(row.get("ราคาต้นทุน (บาท)", 0) or 0),
                            "alert": str(row.get("Alert", "")),
                            "dos": float(row.get("Days of Supply", 0) or 0),
                            "stock": int(row.get("คงเหลือ", 0) or 0),
                            "s30": int(row.get("ขาย 30 วัน", 0) or 0),
                            "s60": int(row.get("ขาย 60 วัน", 0) or 0),
                            "s90": int(row.get("ขาย 90 วัน", 0) or 0),
                        }
                    st.rerun()
            else:
                btn_c, cancel_c = st.columns([3, 1])
                with btn_c:
                    if st.button(f"✅ Confirm สั่งทั้งหมด — {sup} ({len(grp)} รายการ)",
                                 key=f"sup_confirm_{sup_key}",
                                 use_container_width=True, type="primary"):
                        for _, row in grp.iterrows():
                            s = str(row["SKU ID"])
                            qty_val = int(st.session_state.get(f"qty_{s}",
                                          int(row.get("แนะนำสั่ง (หน่วย)", 0) or 0)))
                            st.session_state.confirmed[s] = {
                                "sku_id": s,
                                "sku_name": str(row.get("ชื่อสินค้า", "")),
                                "supplier": sup,
                                "qty": qty_val,
                                "unit_cost": float(row.get("ราคาต้นทุน (บาท)", 0) or 0),
                                "alert": str(row.get("Alert", "")),
                                "dos": float(row.get("Days of Supply", 0) or 0),
                                "stock": int(row.get("คงเหลือ", 0) or 0),
                                "s30": int(row.get("ขาย 30 วัน", 0) or 0),
                                "s60": int(row.get("ขาย 60 วัน", 0) or 0),
                                "s90": int(row.get("ขาย 90 วัน", 0) or 0),
                            }
                        st.rerun()
                with cancel_c:
                    if any_confirmed:
                        if st.button("❌ ยกเลิกทั้งหมด",
                                     key=f"sup_cancel_{sup_key}", use_container_width=True):
                            for s in sup_skus:
                                st.session_state.confirmed.pop(s, None)
                            st.rerun()

            st.markdown("<div style='margin-bottom:16px'></div>", unsafe_allow_html=True)
    else:
        for _, row in order_rows.iterrows():
            render_item(row)


# ── OK / OVERSTOCK ────────────────────────────────────────────────
if len(other_rows) > 0 and any(l in show_levels for l in ["OK", "OVERSTOCK"]):
    with st.expander(f"📋 รายการอื่น (OK / OVERSTOCK) — {len(other_rows)} รายการ"):
        show_cols = ["Alert", "SKU ID", "ชื่อสินค้า", "Supplier",
                     "Days of Supply", "คงเหลือ",
                     "ขาย 15 วัน", "ขาย 30 วัน", "ขาย 60 วัน", "ขาย 90 วัน"]
        show_cols = [c for c in show_cols if c in other_rows.columns]
        st.dataframe(other_rows[show_cols].reset_index(drop=True),
                     use_container_width=True, height=300)


# ══════════════════════════════════════════════════════════════════
#  CONFIRMED SUMMARY + EXPORT
# ══════════════════════════════════════════════════════════════════
st.markdown("---")

if st.session_state.confirmed:
    confirmed_list = [v for v in st.session_state.confirmed.values() if v.get("qty", 0) > 0]
    conf_total = sum(v["qty"] * v["unit_cost"] for v in confirmed_list)
    conf_count = len(confirmed_list)
    sup_count  = len(set(v["supplier"] for v in confirmed_list))

    st.markdown(f"""
<div class="conf-banner">
  <div>
    <div class="cb-title">✅ รายการที่ Confirm แล้ว</div>
    <div class="cb-sub">{conf_count} รายการ จาก {sup_count} ซัพลายเยอร์</div>
  </div>
  <div style="text-align:right">
    <div class="cb-lbl">ยอดสั่งซื้อรวม</div>
    <div class="cb-total">฿{conf_total:,.0f}</div>
  </div>
</div>
""", unsafe_allow_html=True)

    # ตาราง confirmed
    conf_df = pd.DataFrame(confirmed_list)
    conf_df["มูลค่า (บาท)"] = conf_df["qty"] * conf_df["unit_cost"]
    disp = conf_df[["supplier","sku_id","sku_name","qty","unit_cost","มูลค่า (บาท)"]].rename(columns={
        "supplier": "Supplier", "sku_id": "SKU ID", "sku_name": "ชื่อสินค้า",
        "qty": "สั่ง (ชิ้น)", "unit_cost": "ราคา/ชิ้น",
    }).sort_values(["Supplier", "ชื่อสินค้า"])
    st.dataframe(disp, use_container_width=True, hide_index=True)

    # ── Export Excel ─────────────────────────────────────────────
    def build_po_excel(confirmed_list: list) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb   = Workbook()
        wb.remove(wb.active)
        NAVY, GREEN, GOLD, LIGHT = "2F4F8F", "1a5c2a", "FFD700", "EAF4FF"
        thin = Border(
            left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
        )
        today_str = datetime.now().strftime("%d/%m/%Y")

        # summary sheet
        ws_sum = wb.create_sheet("📋 สรุปรวม")
        ws_sum.merge_cells("A1:F1")
        ws_sum["A1"] = f"ใบสั่งซื้อรวม  |  วันที่: {today_str}"
        ws_sum["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
        ws_sum["A1"].fill      = PatternFill("solid", fgColor=GREEN)
        ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[1].height = 30

        hdrs = ["Supplier","SKU ID","ชื่อสินค้า","จำนวนสั่ง (ชิ้น)","ราคา/ชิ้น (บาท)","มูลค่า (บาท)"]
        for ci, h in enumerate(hdrs, 1):
            c = ws_sum.cell(2, ci, h)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin

        by_sup = {}
        for v in sorted(confirmed_list, key=lambda x: (x["supplier"], x["sku_name"])):
            by_sup.setdefault(v["supplier"], []).append(v)

        ri, grand = 3, 0.0
        for sup, items in by_sup.items():
            sup_tot = 0.0
            for item in items:
                val = item["qty"] * item["unit_cost"]
                sup_tot += val; grand += val
                for ci, v in enumerate([sup, item["sku_id"], item["sku_name"],
                                         item["qty"], item["unit_cost"], val], 1):
                    cell = ws_sum.cell(ri, ci, v)
                    cell.border = thin
                    cell.fill   = PatternFill("solid", fgColor=LIGHT)
                    cell.alignment = Alignment(vertical="center",
                                               horizontal="right" if ci >= 4 else "left")
                    if ci in (5, 6): cell.number_format = "#,##0.00"
                    elif ci == 4:    cell.number_format = "#,##0"
                ri += 1
            # subtotal
            for ci in range(1, 7):
                c = ws_sum.cell(ri, ci)
                c.border = thin; c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="D0E8FF")
                c.alignment = Alignment(horizontal="right", vertical="center")
            ws_sum.cell(ri, 3, f"รวม {sup}")
            ws_sum.cell(ri, 4, sum(i["qty"] for i in items)).number_format = "#,##0"
            ws_sum.cell(ri, 6, sup_tot).number_format = "#,##0.00"
            ri += 1

        # grand total
        for ci in range(1, 7):
            c = ws_sum.cell(ri, ci)
            c.border = thin
            c.font = Font(bold=True, size=12, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=GREEN)
            c.alignment = Alignment(horizontal="right", vertical="center")
        ws_sum.cell(ri, 3, "ยอดรวมทั้งหมด")
        ws_sum.cell(ri, 4, sum(v["qty"] for v in confirmed_list)).number_format = "#,##0"
        gc = ws_sum.cell(ri, 6, grand)
        gc.number_format = "#,##0.00"
        gc.font = Font(bold=True, size=12, color=GOLD)
        ws_sum.row_dimensions[ri].height = 28
        for ci, w in enumerate([22,14,36,18,18,18], 1):
            ws_sum.column_dimensions[get_column_letter(ci)].width = w
        ws_sum.freeze_panes = "A3"

        # per-supplier sheets
        for sup, items in by_sup.items():
            sn = sup[:28].replace("/","-").replace("\\","-").replace("*","")
            ws = wb.create_sheet(f"🏭 {sn}")
            ws.merge_cells("A1:F1")
            ws["A1"] = f"ใบสั่งซื้อ: {sup}  |  {today_str}"
            ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
            ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            for ci, h in enumerate(["ลำดับ","SKU ID","ชื่อสินค้า","จำนวนสั่ง (ชิ้น)","ราคา/ชิ้น (บาท)","มูลค่า (บาท)"], 1):
                c = ws.cell(2, ci, h)
                c.font = Font(bold=True, color="FFFFFF", size=10)
                c.fill = PatternFill("solid", fgColor=NAVY)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin

            sup_tot = 0.0
            for i, item in enumerate(sorted(items, key=lambda x: x["sku_name"]), 1):
                val = item["qty"] * item["unit_cost"]; sup_tot += val
                for ci, v in enumerate([i, item["sku_id"], item["sku_name"],
                                         item["qty"], item["unit_cost"], val], 1):
                    c = ws.cell(i+2, ci, v)
                    c.border = thin
                    c.fill   = PatternFill("solid", fgColor="F7FBFF" if i%2==0 else "FFFFFF")
                    c.alignment = Alignment(vertical="center",
                                            horizontal="right" if ci>=4 else ("center" if ci==1 else "left"))
                    if ci in (5,6): c.number_format = "#,##0.00"
                    elif ci==4:     c.number_format = "#,##0"

            tr = len(items)+3
            for ci in range(1, 7):
                c = ws.cell(tr, ci)
                c.border = thin
                c.font = Font(bold=True, size=12, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=NAVY)
                c.alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(tr, 3, "ยอดรวม")
            ws.cell(tr, 4, sum(i["qty"] for i in items)).number_format = "#,##0"
            tc = ws.cell(tr, 6, sup_tot)
            tc.number_format = "#,##0.00"
            tc.font = Font(bold=True, size=13, color=GOLD)
            ws.row_dimensions[tr].height = 26
            for ci, w in enumerate([8,14,36,18,18,18], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.freeze_panes = "A3"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    po_bytes = build_po_excel(confirmed_list)
    fname    = f"PO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    if IS_MOBILE:
        st.download_button(
            "📥 Export PO Excel (แยก Supplier)",
            data=po_bytes, file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", use_container_width=True,
        )
        st.caption(f"จะได้ {sup_count} sheet (แยกซัพลายเยอร์) + 1 sheet สรุปรวม")
    else:
        c1, c2 = st.columns([1, 2])
        with c1:
            st.download_button(
                "📥 Export PO Excel (แยก Supplier)",
                data=po_bytes, file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", use_container_width=True,
            )
        with c2:
            st.info(f"Excel จะมี **{sup_count} sheet** (แยกซัพลายเยอร์) + 1 sheet สรุปรวม")

else:
    st.info("กด **✅ Confirm สั่ง** บนการ์ดสินค้าด้านบน เพื่อเพิ่มรายการและ Export ใบสั่งซื้อ")
