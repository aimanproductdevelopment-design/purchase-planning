"""
APPS — Orchestrator (run_planning.py)
รัน Planning Pipeline ครบสาย:
  Adapter → Data Prep → Forecast → Safety Stock → Reorder → Order Qty → Alert → Excel
"""
import sys
import os
import math
from datetime import date, datetime, timedelta
from collections import defaultdict
from typing import List, Dict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import pandas as pd

from models.config_loader import AppConfig
from models.schema import (
    SkuMaster, SalesRecord, StockRecord, PendingPO, PlanningOutput
)
from adapters.jst import JSTAdapter
from engine.forecast import compute_forecast
from engine.safety_stock import compute_safety_stock
from engine.reorder import compute_reorder
from engine.order_qty import compute_order_qty
from engine.alert import compute_alert, LEVELS


def run(config_path: str = "config.yaml", output_path: str = "output/purchasing_plan.xlsx"):
    print("=" * 60)
    print("  APPS — Aiman Purchase Planning System")
    print(f"  รันวันที่: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    cfg = AppConfig(config_path)
    today = date.today()
    snapshot_at = datetime.now().isoformat(timespec="seconds")

    print("\n[1/5] ดึงข้อมูลจาก JST ERP...")
    adapter = JSTAdapter(cfg)

    skus: List[SkuMaster] = adapter.get_skus()
    print(f"  SKU: {len(skus)} รายการ")

    history_start = today - timedelta(days=430)
    sales: List[SalesRecord] = adapter.get_sales_history(history_start, today)
    print(f"  ยอดขาย: {len(sales):,} records ({history_start} → {today})")

    stocks: List[StockRecord] = adapter.get_stock_on_hand()
    stock_map: Dict[str, StockRecord] = {s.sku_id: s for s in stocks}
    print(f"  สต็อก: {len(stocks)} SKU")

    pos: List[PendingPO] = adapter.get_pending_po()
    po_map: Dict[str, List[PendingPO]] = defaultdict(list)
    for po in pos:
        po_map[po.sku_id].append(po)
    print(f"  PO ค้างรับ: {len(pos)} รายการ")

    print("\n[2/5] เตรียมข้อมูล...")
    sales_by_sku: Dict[str, List[SalesRecord]] = defaultdict(list)
    for s in sales:
        sales_by_sku[s.sku_id].append(s)

    # ดึงยอดขายจริงแต่ละช่วงจากไฟล์ sales_{n}d.xlsx ที่ exportแยกไว้
    print("  ดึงยอดขายแยกช่วง 3/7/15/30/60/90 วันจากไฟล์ JST...")
    sales_window: Dict[str, dict] = adapter.get_sales_windows()
    if not sales_window:
        print("  ⚠️  ไม่พบไฟล์ sales window — ใช้ค่าศูนย์ (รัน jst_exporter เพื่อเติมข้อมูล)")
    else:
        print(f"  ✅ ได้ยอดขายจริง {len(sales_window)} SKU")

    print("\n[3/5] คำนวณ Forecast / Safety Stock / ROP / Order Qty...")
    results: List[PlanningOutput] = []

    for sku in skus:
        if sku.status == "discontinued":
            continue

        lead_time = cfg.lead_time(sku)
        fc = compute_forecast(sku, sales_by_sku.get(sku.sku_id, []), today, lead_time, cfg)
        ss = compute_safety_stock(sku, fc, lead_time, cfg)
        stock = stock_map.get(sku.sku_id) or StockRecord(sku_id=sku.sku_id, on_hand=0.0)
        reorder = compute_reorder(sku.sku_id, stock, po_map.get(sku.sku_id, []), fc, ss, lead_time)
        order = compute_order_qty(sku, fc, reorder, lead_time, cfg,
                                   sales_window=sales_window.get(sku.sku_id))
        alert = compute_alert(sku.sku_id, reorder, ss, lead_time, cfg, fc=fc)

        results.append(PlanningOutput(
            sku=sku, forecast=fc, safety_stock=ss,
            reorder=reorder, order_qty=order, alert=alert,
            lead_time_used=lead_time, snapshot_at=snapshot_at,
        ))

    print(f"  คำนวณเสร็จ {len(results)} SKU")

    print("\n[4/5] สรุป Alert:")
    level_count = defaultdict(int)
    for r in results:
        level_count[r.alert.level] += 1
    for lvl in LEVELS:
        if level_count[lvl]:
            print(f"  {lvl}: {level_count[lvl]} SKU")

    print(f"\n[5/5] บันทึก Excel → {output_path}")
    _export_excel(results, output_path, snapshot_at, sales_window)

    print("\n[6/6] LINE Notify...")
    try:
        from notify.line_notify import send_summary
        send_summary(results, output_path=output_path)
    except Exception as e:
        print(f"   ⚠️  LINE Notify ข้าม: {e}")

    print("\n✅ เสร็จสิ้น!")
    return results


# ── สีตามระดับ Alert ─────────────────────────────────────────────
_ALERT_BG = {
    "CRITICAL":  "FFCCCC",
    "WARNING":   "FFE5CC",
    "WATCH":     "FFFACC",
    "OK":        "CCFFCC",
    "OVERSTOCK": "E0E0E0",
}
_HEADER_BG = "2F4F8F"
_HEADER_FG = "FFFFFF"


def _make_row(r, sales_window: dict = None) -> dict:
    sw = (sales_window or {}).get(r.sku.sku_id, {})
    return {
        "Alert":                r.alert.level,
        "SKU ID":               r.sku.sku_id,
        "ชื่อสินค้า":           r.sku.sku_name,
        "หมวด":                 r.sku.category,
        "Supplier":             r.sku.supplier_id,
        "Lead Time (วัน)":      r.lead_time_used,
        "คงเหลือ":              round(r.reorder.available, 0),
        "ขาย 3 วัน":            sw.get("s3", 0),
        "ขาย 7 วัน":            sw.get("s7", 0),
        "ขาย 15 วัน":           sw.get("s15", 0),
        "ขาย 30 วัน":           sw.get("s30", 0),
        "ขาย 60 วัน":           sw.get("s60", 0),
        "ขาย 90 วัน":           sw.get("s90", 0),
        "PO ค้างรับ":           round(r.reorder.on_order, 0),
        "Inv. Position":        round(r.reorder.inventory_position, 0),
        "Forecast/วัน":         round(r.forecast.daily, 2),
        "Days of Supply":       round(r.reorder.days_of_supply, 1),
        "Safety Stock":         round(r.safety_stock.safety_stock, 0),
        "Reorder Point":        round(r.reorder.rop, 0),
        "ต้องสั่ง?":            "ใช่" if r.reorder.need_to_order else "",
        "ราคาต้นทุน (บาท)":     round(r.sku.unit_cost, 2),
        "แนะนำสั่ง (หน่วย)":    int(r.order_qty.suggested_qty) if r.order_qty.suggested_qty > 0 else "",
        "มูลค่าสั่ง (บาท)":     round(r.order_qty.suggested_qty * r.sku.unit_cost, 2)
                                 if r.order_qty.suggested_qty > 0 else "",
        "หมายเหตุ":             r.alert.message or r.order_qty.reason,
    }


def _style_sheet(ws, df: pd.DataFrame, alert_col: str = "Alert"):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_font = Font(color=_HEADER_FG, bold=True, size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    alert_col_idx = None
    for i, col in enumerate(df.columns, 1):
        if col == alert_col:
            alert_col_idx = i
            break

    if alert_col_idx:
        for row_idx in range(2, ws.max_row + 1):
            level = ws.cell(row_idx, alert_col_idx).value or ""
            bg = _ALERT_BG.get(level, "FFFFFF")
            fill = PatternFill("solid", fgColor=bg)
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row_idx, col_idx)
                cell.fill = fill
                cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col in enumerate(df.columns, 1):
        letter = get_column_letter(col_idx)
        max_len = len(str(col))
        for cell in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            max_len = max(max_len, len(str(cell[0].value or "")))
        ws.column_dimensions[letter].width = min(max_len + 3, 45)

    ws.row_dimensions[1].height = 28


def _export_excel(results: List[PlanningOutput], path: str, snapshot_at: str, sales_window: dict = None):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    level_order = {l: i for i, l in enumerate(LEVELS)}
    sorted_results = sorted(
        results,
        key=lambda r: (level_order.get(r.alert.level, 99), r.reorder.days_of_supply)
    )

    all_rows    = [_make_row(r, sales_window) for r in sorted_results]
    critical    = [_make_row(r, sales_window) for r in sorted_results if r.alert.level == "CRITICAL"]
    watch_warn  = [_make_row(r, sales_window) for r in sorted_results if r.alert.level in ("WARNING", "WATCH")]
    normal_over = [_make_row(r, sales_window) for r in sorted_results if r.alert.level in ("OK", "OVERSTOCK")]

    audit_rows = []
    for r in sorted_results:
        audit_rows.append({
            "SKU ID":           r.sku.sku_id,
            "ชื่อสินค้า":       r.sku.sku_name,
            "snapshot_at":      snapshot_at,
            "lead_time_used":   r.lead_time_used,
            "weighted_avg":     round(r.forecast.weighted_add, 4),
            "seasonal_factor":  round(r.forecast.seasonal_factor, 4),
            "std_dev_daily":    round(r.forecast.std_dev_daily, 4),
            "forecast_daily":   round(r.forecast.daily, 4),
            "forecast_lt":      round(r.forecast.lead_time_forecast, 2),
            "ss_mode":          r.safety_stock.mode_used,
            "safety_stock":     round(r.safety_stock.safety_stock, 0),
            "rop":              round(r.reorder.rop, 0),
            "available":        round(r.reorder.available, 0),
            "on_order":         round(r.reorder.on_order, 0),
            "inv_position":     round(r.reorder.inventory_position, 0),
            "days_of_supply":   round(r.reorder.days_of_supply, 1),
            "suggested_qty":    int(r.order_qty.suggested_qty),
            "alert_level":      r.alert.level,
            "flags":            ", ".join(r.forecast.flags),
        })

    sheets = [
        ("สั่งด่วน (CRITICAL)",  critical),
        ("เฝ้าระวัง",            watch_warn),
        ("ปกติ+OVERSTOCK",       normal_over),
        ("ทั้งหมด",              all_rows),
    ]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, rows in sheets:
            df = pd.DataFrame(rows) if rows else pd.DataFrame(
                columns=list(_make_row(results[0], sales_window).keys()) if results else [])
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_sheet(writer.sheets[sheet_name], df)

        df_audit = pd.DataFrame(audit_rows)
        df_audit.to_excel(writer, sheet_name="Audit", index=False)
        _style_sheet(writer.sheets["Audit"], df_audit, alert_col="alert_level")


if __name__ == "__main__":
    cfg_path = sys.argv[1] if len(sys.argv) > 1 else "config.yaml"
    run(config_path=cfg_path)
